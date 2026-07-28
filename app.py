import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- GENEL STİL TANIMLARI ---
DARK_BLUE = "1F4E78"
LIGHT_BLUE = "D9E1F2"
GRAY_HEADER = "595959"
GRAY_LIGHT = "F2F2F2"
WHITE = "FFFFFF"

font_header = Font(name="Calibri", size=11, bold=True, color=WHITE)
font_bold = Font(name="Calibri", size=11, bold=True)

fill_dark_blue = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
fill_light_blue = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
fill_gray_header = PatternFill(start_color=GRAY_HEADER, end_color=GRAY_HEADER, fill_type="solid")

thin_border_side = Side(border_style="thin", color="D9D9D9")
thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

align_center = Alignment(horizontal="center", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")


def apply_autofit_and_styles(ws):
    """Sütun genişliklerini otomatik ayarlar ve hücre sınırlarını çizip ızgara çizgilerini aktif eder."""
    ws.views.sheetView[0].showGridLines = True
    
    # Tüm hücrelere border ve varsayılan hiza uygula
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if not cell.border.left.style:
                cell.border = thin_border

    # Sütun Genişlikleri
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Merged (birleştirilmiş) hücrelerden kaynaklı hatayı önlemek için
            if type(cell).__name__ == 'MergedCell':
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 5, 14)


# ==========================================
# 1. DOSYA: AT ZİMMET İZLEME.xlsx
# ==========================================
def generate_at_zimmet_izleme(df_raw, filename="AT_ZIMMET_IZLEME.xlsx"):
    wb = openpyxl.Workbook()
    
    # 1.1 Şube Performansı ve Kanal Dağılımı
    ws1 = wb.active
    ws1.title = "Şube Performansı"
    
    ws1.merge_cells("A1:E1")
    ws1["A1"] = "ŞUBE PERFORMANSI VE KANAL DAĞILIMI"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    ws1["A1"].fill = fill_dark_blue
    ws1["A1"].alignment = align_center

    headers1 = ["Kanal / Metrik", "Toplam Adet", "Teslim Adet", "Kullanıcı İade", "Başarı Oranı (%)"]
    for col_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_gray_header
        cell.alignment = align_center

    total_zimmet = len(df_raw) if not df_raw.empty else 0
    total_teslim = len(df_raw[df_raw['Durum'] == 'Teslim Edildi']) if ('Durum' in df_raw.columns and not df_raw.empty) else 0
    total_iade = len(df_raw[df_raw['Durum'] == 'İade']) if ('Durum' in df_raw.columns and not df_raw.empty) else 0
    
    metrics = [
        ["Saha / Kurye Dağıtımı", total_zimmet, total_teslim, total_iade, "=IF(B4>0, C4/B4, 0)"],
        ["Genel Toplam", "=SUM(B4:B4)", "=SUM(C4:C4)", "=SUM(D4:D4)", "=IF(B5>0, C5/B5, 0)"]
    ]
    
    for r_idx, row_data in enumerate(metrics, start=4):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = align_right if c_idx > 1 else align_left
            if r_idx == 5:
                cell.font = font_bold
                cell.fill = fill_light_blue
            if c_idx == 5:
                cell.number_format = "0.0%"

    # 1.2 Personel Bazlı Karşılaştırmalı Teslimat Grafiği (Veri Tabanı)
    ws2 = wb.create_sheet(title="Personel Grafiği Verileri")
    ws2.append(["Personel", "Teslimat Adedi", "Hedef"])
    if 'Personel' in df_raw.columns and not df_raw.empty:
        p_counts = df_raw['Personel'].value_counts()
        for p_name, count in p_counts.items():
            ws2.append([p_name, count, 100])

    # 1.3 Genel Durum ve Performans
    ws3 = wb.create_sheet(title="Genel Durum")
    ws3.append(["Metrik", "Değer"])
    ws3.append(["Toplam Dağıtıma Çıkan", total_zimmet])
    ws3.append(["Toplam Teslim Edilen", total_teslim])
    ws3.append(["Genel Başarı Yüzdesi", f"={(total_teslim/total_zimmet):.1%}" if total_zimmet > 0 else "0.0%"])

    # 1.4 Personel Zimmet & Teslim Özeti Tablosu
    ws4 = wb.create_sheet(title="Zimmet & Teslim Özeti")
    ws4.append(["Personel Adı", "Zimmet Edilen Paket", "Teslim Edilen", "Kalan Paket", "Başarı Oranı"])
    if 'Personel' in df_raw.columns and not df_raw.empty:
        summary = df_raw.groupby('Personel').size().reset_index(name='Zimmet')
        for idx, row in summary.iterrows():
            r = idx + 2
            teslim_cnt = len(df_raw[(df_raw['Personel'] == row['Personel']) & (df_raw['Durum'] == 'Teslim Edildi')])
            ws4.append([row['Personel'], row['Zimmet'], teslim_cnt, f"=B{r}-C{r}", f"=IF(B{r}>0, C{r}/B{r}, 0)"])
            ws4.cell(row=r, column=5).number_format = "0.0%"

    for ws in wb.worksheets:
        apply_autofit_and_styles(ws)
        
    wb.save(filename)


# ==========================================
# 2. DOSYA: PERSONEL HESAP ALIMI EKRANI.xlsx
# ==========================================
def generate_personel_hesap_alimi(df_raw, filename="PERSONEL_HESAP_ALIMI_EKRANI.xlsx"):
    wb = openpyxl.Workbook()
    
    # 2.1 Personel Hesap Alımı Ekranı
    ws1 = wb.active
    ws1.title = "Hesap Alımı Ekranı"
    
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "PERSONEL HESAP ALIMI EKRANI"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    ws1["A1"].fill = fill_dark_blue
    ws1["A1"].alignment = align_center

    headers1 = ["Sıra", "Personel Adı Soyadı", "Nakit Tahsilat", "POS Tahsilat", "Toplam Tahsilat", "Teslim Edilen Adet", "Durum"]
    for c_idx, h in enumerate(headers1, start=1):
        cell = ws1.cell(row=3, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_gray_header
        cell.alignment = align_center

    start_row = 4
    personeller = df_raw['Personel'].unique() if ('Personel' in df_raw.columns and not df_raw.empty) else []
    
    if len(personeller) > 0:
        for idx, p in enumerate(personeller, start=1):
            r = start_row + idx - 1
            ws1.cell(row=r, column=1, value=idx).alignment = align_center
            ws1.cell(row=r, column=2, value=p).alignment = align_left
            ws1.cell(row=r, column=3, value=0.0).number_format = '#,##0.00 TL'
            ws1.cell(row=r, column=4, value=0.0).number_format = '#,##0.00 TL'
            
            # Toplam Tahsilat = Nakit + POS
            cell_tot = ws1.cell(row=r, column=5, value=f"=C{r}+D{r}")
            cell_tot.number_format = '#,##0.00 TL'
            cell_tot.font = font_bold
            
            p_teslim = len(df_raw[(df_raw['Personel'] == p) & (df_raw['Durum'] == 'Teslim Edildi')]) if 'Durum' in df_raw.columns else 0
            ws1.cell(row=r, column=6, value=p_teslim).alignment = align_center
            ws1.cell(row=r, column=7, value="Tamamlandı").alignment = align_center

        end_row = start_row + len(personeller) - 1
    else:
        # Boş veri durumu için varsayılan tek satır
        end_row = start_row
        ws1.cell(row=end_row, column=2, value="Veri Bulunamadı")

    # İSTEĞİNİZ: Tüm Personellerin Toplam Tahsilat Değerleri Toplamı
    tot_row = end_row + 2
    ws1.cell(row=tot_row, column=2, value="PERSONELLER TOPLAM TAHSİLAT HESABI:").font = font_bold
    ws1.cell(row=tot_row, column=2).alignment = align_right
    
    sum_cell = ws1.cell(row=tot_row, column=5, value=f"=SUM(E{start_row}:E{end_row})")
    sum_cell.font = Font(name="Calibri", size=11, bold=True, color="006100")
    sum_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    sum_cell.number_format = '#,##0.00 TL'

    # 2.2 Tüm Personellerin Hesap Alımı Özeti Tablosu
    ws2 = wb.create_sheet(title="Hesap Alımı Özeti")
    ws2.append(["Personel", "Nakit", "POS", "Genel Toplam"])
    if len(personeller) > 0:
        for idx, p in enumerate(personeller, start=2):
            ws2.append([p, 0.0, 0.0, f"=B{idx}+C{idx}"])

    for ws in wb.worksheets:
        apply_autofit_and_styles(ws)

    wb.save(filename)


# ==========================================
# 3. DOSYA: F4 ÖDEME LİSTESİ.xlsx
# ==========================================
def generate_f4_odeme_listesi(df_raw, filename="F4_ODEME_LISTESI.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "F4 Ödeme Listesi"

    ws.merge_cells("A1:F1")
    ws["A1"] = "F4 ÖDEME LİSTESİ (PERSONEL BAZLI SÜZGEÇ)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=WHITE)
    ws["A1"].fill = fill_dark_blue
    ws["A1"].alignment = align_center

    headers = ["Takip No", "Personel", "Alıcı Adı", "Ödeme Tipi", "Tutar", "Ödeme Durumu"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_gray_header
        cell.alignment = align_center

    if not df_raw.empty:
        for r_idx, row in df_raw.reset_index(drop=True).iterrows():
            row_num = r_idx + 4
            ws.cell(row=row_num, column=1, value=row.get('Takip No', f'TKP{r_idx+1}')).alignment = align_center
            ws.cell(row=row_num, column=2, value=row.get('Personel', '')).alignment = align_left
            ws.cell(row=row_num, column=3, value=row.get('Alıcı', '')).alignment = align_left
            ws.cell(row=row_num, column=4, value=row.get('Ödeme Tipi', 'Nakit')).alignment = align_center
            
            c_tutar = ws.cell(row=row_num, column=5, value=float(row.get('Tutar', 0.0)))
            c_tutar.number_format = '#,##0.00 TL'
            c_tutar.alignment = align_right
            
            ws.cell(row=row_num, column=6, value=row.get('Ödeme Durumu', 'Tahsil Edildi')).alignment = align_center

    max_row = max(ws.max_row, 4)
    ws.auto_filter.ref = f"A3:F{max_row}"

    apply_autofit_and_styles(ws)
    wb.save(filename)


# ==========================================
# TEST İŞLEMİ
# ==========================================
if __name__ == "__main__":
    raw_data = {
        'Takip No': ['TK1001', 'TK1002', 'TK1003', 'TK1004'],
        'Personel': ['Ahmet Yılmaz', 'Mehmet Demir', 'Ahmet Yılmaz', 'Ayşe Kaya'],
        'Durum': ['Teslim Edildi', 'Teslim Edildi', 'İade', 'Teslim Edildi'],
        'Alıcı': ['Ali Can', 'Veli Han', 'Ayşe Tan', 'Fatma Şen'],
        'Ödeme Tipi': ['Nakit', 'POS', 'Nakit', 'POS'],
        'Tutar': [150.00, 320.50, 0.00, 450.00],
        'Ödeme Durumu': ['Tahsil Edildi', 'Tahsil Edildi', 'İptal', 'Tahsil Edildi']
    }
    df = pd.DataFrame(raw_data)

    generate_at_zimmet_izleme(df)
    generate_personel_hesap_alimi(df)
    generate_f4_odeme_listesi(df)
    print("İşlem başarıyla tamamlandı. 3 ayrı dosya oluşturuldu.")
